#!/usr/bin/env python
# coding: utf-8

# In[10]:


from openpyxl import load_workbook
import shutil
import arcgis
from arcgis.gis import GIS
import re
import csv
import pandas as pd
import os
# Define variables
portalURL = r'https://gissp.bracell.com/portal/'
username = "Qualidade_Florestal"
password = "Qualidade@24"
survey_item_id = "70e124db3aa34a8c816bbfa9dbee237d"
save_path = r"F:\Qualidade_Florestal\03- ADMINISTRATIVO\2023\06- COLABORADORES\Gabriel\1 - Processamento de dados"


# In[ ]:


keep_org_item = False

# Connect to GIS Portal and identify Survey form
gis = GIS(portalURL, username, password)
survey_by_id = gis.content.get(survey_item_id)
print(survey_by_id.type)
survey_by_id

# Download service
rel_fs = survey_by_id.related_items('Survey2Service', 'forward')[0]
rel_fs

item_excel = rel_fs.export(title=survey_by_id.title, export_format='Excel')
item_excel.download(save_path=save_path)
if not bool(keep_org_item):
    item_excel.delete(force=True)


# In[ ]:


import pandas as pd
from openpyxl import load_workbook

# Carregar o arquivo Excel existente
path_arquivo = os.path.join(save_path, "Avaliação_de_Sobrevivência_30_e_90.xlsx")

# Ler a aba específica em um DataFrame
nome_aba = "Formulario_2"
dados = pd.read_excel(path_arquivo, sheet_name=nome_aba)

# Aplicar as modificações desejadas no DataFrame
dados['talhao'] = dados['talhao'].fillna('').astype(str).str.zfill(3)
dados['id_fazenda'] = dados['id_fazenda'].fillna('').astype(str).str.zfill(4)
# Salvar o DataFrame na aba específica mantendo as outras abas
with pd.ExcelWriter(path_arquivo, mode='a', engine='openpyxl') as writer:
    if nome_aba in writer.book.sheetnames:
        writer.book.remove(writer.book[nome_aba])
    dados.to_excel(writer, sheet_name=nome_aba, index=False)

# In[ ]:


destino = (r"F:\Qualidade_Florestal\02- MATO GROSSO DO SUL\08- Planejamento e Controle 2°nível\2023\02 - Programação Sobrevivência\Apontamentos\Avaliação_de_Sobrevivência_30_e_90.xlsx")


# In[ ]:


shutil.copy(path_arquivo, destino)
